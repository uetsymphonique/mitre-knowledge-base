import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_COLUMNS = ["Summary", "Technique", "Procedures", "Description"]


def normalize_text(value: object, joiner: str) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    # Split on any newline variant and strip each part
    parts = [seg.strip() for seg in text.replace("\r", "\n").split("\n")]
    parts = [p for p in parts if p]
    if not parts:
        return ""
    # Join with requested delimiter and collapse any leftover whitespace
    normalized = f" {joiner} ".join(parts)
    return normalized.strip()


def find_header_indices(header_cells: list[str]) -> dict[str, int]:
    indices: dict[str, int] = {}
    normalized = {str(name).strip().lower(): idx for idx, name in enumerate(header_cells)}
    for col in EXPECTED_COLUMNS:
        key = col.lower()
        if key not in normalized:
            raise ValueError(f"Required column '{col}' not found in header: {header_cells}")
        indices[col] = normalized[key]
    return indices


def main():
    parser = argparse.ArgumentParser(
        description="Normalize KB TTPs XLSX (Summary, Technique, Procedures, Description) to CSV"
    )
    parser.add_argument("input", help="Path to input .xlsx file")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name (default: first sheet)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Path to output file (default: alongside input with extension per --format)",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "json", "oneline"],
        default="csv",
        help="Output format (default: csv). 'oneline' writes NDJSON (one JSON object per line) to .txt",
    )
    parser.add_argument(
        "--tech2tac",
        default=None,
        help="Optional path to tech2tac JSON to append Tactics column based on Technique IDs",
    )
    parser.add_argument(
        "--seperate-by-technique",
        action="store_true",
        help="Write one file per technique named <technique>_<timestamp> in --outdir",
    )
    parser.add_argument(
        "--seperate-by-tactic",
        action="store_true",
        help="Write one file per tactic named <tactic>_<timestamp> in --outdir (requires --tech2tac)",
    )
    parser.add_argument(
        "--outdir",
        default=None,
        help="Output directory when using --seperate-by-technique or --seperate-by-tactic (default: alongside input)",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=None,
        help="Split outputs into chunks of at most N records (works with -o, --seperate-by-technique, and --seperate-by-tactic)",
    )
    args = parser.parse_args()

    # Validation
    if args.seperate_by_technique and args.seperate_by_tactic:
        raise ValueError("Cannot use both --seperate-by-technique and --seperate-by-tactic together")
    
    if args.seperate_by_tactic and not args.tech2tac:
        raise ValueError("--seperate-by-tactic requires --tech2tac mapping file")

    in_path = Path(args.input)
    if not in_path.exists():
        raise FileNotFoundError(f"Input file not found: {in_path}")

    # Choose default extension based on format
    default_ext = ".json" if args.format == "json" else (".txt" if args.format == "oneline" else ".csv")
    out_path = Path(args.output) if args.output else in_path.with_suffix(default_ext)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=str(in_path), read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("Input sheet is empty")

    header_cells = [str(h) if h is not None else "" for h in header]
    col_idx = find_header_indices(header_cells)

    # Optional mapping of technique ID => [tactic IDs]
    id_to_tactics: dict[str, list[str]] | None = None
    if args.tech2tac:
        import json as _json
        map_path = Path(args.tech2tac)
        if not map_path.exists():
            raise FileNotFoundError(f"tech2tac mapping not found: {map_path}")
        id_to_tactics = _json.loads(map_path.read_text(encoding="utf-8"))

    # Collect normalized rows first (useful for both formats)
    records = []
    for row in rows_iter:
        cells = list(row)
        def get(col: str):
            idx = col_idx[col]
            return cells[idx] if idx < len(cells) else None

        summary = normalize_text(get("Summary"), ".")
        technique = normalize_text(get("Technique"), ".")
        procedures = normalize_text(get("Procedures"), ";")
        description = normalize_text(get("Description"), ".")

        # Build Behavior: concatenate non-empty segments with labels
        segments = []
        if summary:
            segments.append(f"Summary: {summary}.")
        if description:
            segments.append(f"Description: {description}.")
        if procedures:
            segments.append(f"Procedures: {procedures}.")
        behavior = " ".join(segments).strip()

        record = {
            "Technique": technique,
            "Behavior": behavior,
        }

        # Append Tactics via mapping if provided
        if id_to_tactics is not None:
            import re as _re
            # Extract all technique IDs from the Technique field
            tech_ids = _re.findall(r"T\d{4}(?:\.\d{3})?", technique)
            tactics_set: set[str] = set()
            for tid in tech_ids:
                key = tid.split(".")[0]
                if key in id_to_tactics and id_to_tactics[key]:
                    for tac in id_to_tactics[key]:
                        tactics_set.add(str(tac))
            tactics_list = sorted(tactics_set)
            record["Tactics"] = "; ".join(tactics_list) if args.format == "csv" else tactics_list

        records.append(record)

    # Separate into per-technique or per-tactic files if requested
    if args.seperate_by_technique or args.seperate_by_tactic:
        from datetime import datetime

        def sanitize_filename(name: str) -> str:
            safe = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in name)
            return safe.strip("._") or "unknown"

        groups: dict[str, list[dict]] = {}
        if args.seperate_by_technique:
            # Group by technique
            for rec in records:
                tech_key = rec.get("Technique", "").strip() or "unknown"
                groups.setdefault(tech_key, []).append(rec)
        else:  # args.seperate_by_tactic
            # Group by individual tactics
            for rec in records:
                tactics_field = rec.get("Tactics", "")
                if args.format == "csv":
                    # Tactics is a string like "TA0001 - Initial Access; TA0002 - Execution"
                    tactic_names = [t.strip() for t in str(tactics_field).split(";") if t.strip()]
                else:
                    # Tactics is a list
                    tactic_names = tactics_field if isinstance(tactics_field, list) else []
                
                if not tactic_names:
                    # Records without tactics go to "unknown" group
                    groups.setdefault("unknown", []).append(rec)
                else:
                    # Add record to each tactic group
                    for tactic in tactic_names:
                        groups.setdefault(tactic, []).append(rec)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_outdir = Path(args.outdir) if args.outdir else in_path.parent
        base_outdir.mkdir(parents=True, exist_ok=True)

        written_files = 0
        for group_key, recs in groups.items():
            ext = 'json' if args.format == 'json' else ('txt' if args.format == 'oneline' else 'csv')
            base_name = f"{sanitize_filename(group_key)}_{ts}"
            chunk_size = args.chunk_size if args.chunk_size and args.chunk_size > 0 else None

            def write_one(path: Path, subset: list[dict]):
                nonlocal written_files
                if args.format == "json":
                    import json
                    path.write_text(json.dumps(subset, ensure_ascii=False, indent=2), encoding="utf-8")
                elif args.format == "oneline":
                    import json
                    lines = [json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in subset]
                    path.write_text("\n".join(lines), encoding="utf-8")
                else:
                    with path.open("w", encoding="utf-8", newline="") as f:
                        writer = csv.writer(f)
                        headers = ["Technique", "Behavior"] + (["Tactics"] if id_to_tactics is not None else [])
                        writer.writerow(headers)
                        for r in subset:
                            row_vals = [r["Technique"], r["Behavior"]]
                            if id_to_tactics is not None:
                                row_vals.append(r.get("Tactics", ""))
                            writer.writerow(row_vals)
                written_files += 1

            if chunk_size:
                for i in range(0, len(recs), chunk_size):
                    part_idx = i // chunk_size + 1
                    fpath = base_outdir / f"{base_name}.part{part_idx}.{ext}"
                    write_one(fpath, recs[i : i + chunk_size])
            else:
                fpath = base_outdir / f"{base_name}.{ext}"
                write_one(fpath, recs)
        print(f"Saved {written_files} files to {base_outdir}")
    else:
        chunk_size = args.chunk_size if args.chunk_size and args.chunk_size > 0 else None

        if args.format == "json":
            import json
            if chunk_size:
                base = out_path.parent / out_path.stem
                for i in range(0, len(records), chunk_size):
                    part_idx = i // chunk_size + 1
                    part_path = Path(f"{base}.part{part_idx}.json")
                    part_path.write_text(json.dumps(records[i : i + chunk_size], ensure_ascii=False, indent=2), encoding="utf-8")
            else:
                out_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        elif args.format == "oneline":
            import json
            lines = [json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in records]
            # ensure txt extension
            if out_path.suffix.lower() != ".txt":
                out_path = out_path.with_suffix(".txt")
            if chunk_size:
                base = out_path.parent / out_path.stem
                for i in range(0, len(lines), chunk_size):
                    part_idx = i // chunk_size + 1
                    part_path = Path(f"{base}.part{part_idx}.txt")
                    part_path.write_text("\n".join(lines[i : i + chunk_size]), encoding="utf-8")
            else:
                out_path.write_text("\n".join(lines), encoding="utf-8")
        else:
            if chunk_size:
                base = out_path.parent / out_path.stem
                for i in range(0, len(records), chunk_size):
                    part_idx = i // chunk_size + 1
                    part_path = Path(f"{base}.part{part_idx}.csv")
                    with part_path.open("w", encoding="utf-8", newline="") as f:
                        writer = csv.writer(f)
                        headers = ["Technique", "Behavior"] + (["Tactics"] if id_to_tactics is not None else [])
                        writer.writerow(headers)
                        for rec in records[i : i + chunk_size]:
                            row_vals = [rec["Technique"], rec["Behavior"]]
                            if id_to_tactics is not None:
                                row_vals.append(rec.get("Tactics", ""))
                            writer.writerow(row_vals)
            else:
                with out_path.open("w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    headers = ["Technique", "Behavior"] + (["Tactics"] if id_to_tactics is not None else [])
                    writer.writerow(headers)
                    for rec in records:
                        row_vals = [rec["Technique"], rec["Behavior"]]
                        if id_to_tactics is not None:
                            row_vals.append(rec.get("Tactics", ""))
                            
                        writer.writerow(row_vals)

    wb.close()
    if not args.seperate_by_technique and not args.seperate_by_tactic:
        print(f"Saved normalized {args.format.upper()} to {out_path}")


if __name__ == "__main__":
    main()


